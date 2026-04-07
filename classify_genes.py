#!/usr/bin/env python3
"""
classify_genes.py — E. coli Gene GO/KEGG/EcoCyc Classifier
Organism  : Escherichia coli K-12 MG1655
APIs used : EcoCyc/BioCyc  https://websvc.biocyc.org  (primary)
            UniProt REST    https://rest.uniprot.org   (secondary)
            KEGG REST       https://rest.kegg.jp       (tertiary)

EcoCyc credentials — set in a .env file next to this script:
    ECOCYC_EMAIL=you@example.com
    ECOCYC_PASSWORD=yourpassword

All grouping is derived from fetched EcoCyc pathway hierarchy.
Only gene names are hardcoded; everything else is resolved at runtime.
"""
from __future__ import annotations
import os, re, sys, time
import xml.etree.ElementTree as ET
from collections import Counter
from dataclasses import dataclass, field

sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]

try:
    import requests
except ImportError:
    sys.exit("pip install requests")


def _load_env_file() -> None:
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.isfile(env_path):
        return
    with open(env_path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            os.environ.setdefault(key.strip(), value.strip())

_load_env_file()

# ── Organism constants ─────────────────────────────────────────────────────────
ORGANISM_FULL  = "Escherichia coli K-12 MG1655"
ORGANISM_SHORT = "E. coli K-12"
ORGANISM_TAXID = "83333"
ORGANISM_KEGG  = "eco"
ECOCYC_ORG     = "ECOLI"

# ── Gene list — only names, nothing else ──────────────────────────────────────
GENE_NAMES: list[str] = [
    "ftsZ", "ftsI", "mrdA", "mrcB", "mrcA", "murA", "murC",
    "lpxA", "lpxC", "lptA", "lptC", "msbA",
    "folA", "folP", "dnaE", "dnaB", "gyrA", "gyrB", "parC", "parE",
    "rpoA", "rpoB", "rpsA", "rpsL", "rplA", "rplC", "secA", "secY",
]

# ── Display palette — no biological meaning ────────────────────────────────────
_PALETTE = [
    "AED6F1", "A9DFBF", "A3E4D7", "F9E79F", "F5CBA7",
    "D2B4DE", "FAD7A0", "FADBD8", "D5DBDB", "ABEBC6",
    "F1948A", "85C1E9", "82E0AA", "F8C471", "C39BD3",
    "76D7C4", "F0B27A", "EC7063", "5DADE2", "45B39D",
]

# KEGG global overview / resistance map IDs to skip
_KEGG_GENERIC = {
    "eco01100", "eco01110", "eco01120", "eco01130",
    "eco01200", "eco01210", "eco01220", "eco01230",
    "eco01501", "eco01502",
}

# Pathway class names too broad to be useful as group labels.
# _pathway_class stops climbing when it finds a name NOT in this set.
_GENERIC_PATHWAY_ROOTS = {
    "biosynthesis", "degradation", "energy metabolism",
    "detoxification", "activation/inactivation/interconversion",
    "macromolecule modification",
}


# ── Data model ─────────────────────────────────────────────────────────────────
@dataclass
class Gene:
    name:             str
    ecocyc_id:        str = ""
    full_name:        str = ""
    go_bp:            dict[str, str] = field(default_factory=dict)  # GO-ID → term
    go_bp_votes:      Counter = field(default_factory=Counter)       # GO-ID → count
    go_mf:            dict[str, str] = field(default_factory=dict)
    go_cc:            dict[str, str] = field(default_factory=dict)
    ecocyc_pathways:  dict[str, str] = field(default_factory=dict)  # frameid → name
    ecocyc_class:     str = ""
    molecular_weight_kd: float | None = None
    accession:        str = ""
    kegg_id:          str = ""
    kegg_pathways:    dict[str, str] = field(default_factory=dict)
    group:            str = "Unclassified"

# ── HTTP / rate-limit helpers ──────────────────────────────────────────────────
_ECOCYC_BASE = "https://websvc.biocyc.org"
_UNIPROT_BASE = "https://rest.uniprot.org/uniprotkb"
_KEGG_BASE    = "https://rest.kegg.jp"
_DELAY        = {"ecocyc": 1.1, "uniprot": 0.15, "kegg": 0.4}

_SESSION = requests.Session()
_SESSION.headers["User-Agent"] = "classify_genes.py/2.0 (research script)"


def login_ecocyc() -> bool:
    email    = os.environ.get("ECOCYC_EMAIL", "")
    password = os.environ.get("ECOCYC_PASSWORD", "")
    if not email or not password:
        print("  [EcoCyc] credentials not set — check your .env file.", file=sys.stderr)
        return False
    print(f"  [EcoCyc] logging in as {email} ...", end=" ", flush=True)
    try:
        r = _SESSION.post(
            f"{_ECOCYC_BASE}/credentials/login/",
            data={"email": email, "password": password},
            timeout=20,
        )
        print(f"HTTP {r.status_code}", end=" ", flush=True)
    except Exception as exc:
        print(f"\n  [EcoCyc] login error: {exc}", file=sys.stderr)
        return False
    test = _SESSION.get(f"{_ECOCYC_BASE}/getxml?{ECOCYC_ORG}:EG10347", timeout=20)
    time.sleep(_DELAY["ecocyc"])
    if test.status_code == 200 and test.text.strip().startswith("<?xml"):
        print("— session OK")
        return True
    print(
        f"\n  [EcoCyc] session check failed (HTTP {test.status_code})."
        f"\n  Login body: {r.text[:200]!r}"
        f"\n  Test body:  {test.text[:200]!r}",
        file=sys.stderr,
    )
    return False


def _http_get(url: str, params: dict | None = None, source: str = "ecocyc") -> requests.Response:
    r = _SESSION.get(url, params=params, timeout=30)
    r.raise_for_status()
    time.sleep(_DELAY[source])
    return r


# ── XML helpers (namespace-tolerant) ──────────────────────────────────────────
def _local(tag: str) -> str:
    return tag.split("}")[-1] if "}" in tag else tag

def _find(root: ET.Element, *local_tags: str) -> ET.Element | None:
    for el in root.iter():
        if _local(el.tag) in local_tags:
            return el
    return None

def _findall(root: ET.Element, local_tag: str) -> list[ET.Element]:
    return [el for el in root.iter() if _local(el.tag) == local_tag]

def _text(root: ET.Element, *local_tags: str) -> str:
    el = _find(root, *local_tags)
    return (el.text or "").strip() if el is not None else ""

def _parse(xml_text: str) -> ET.Element | None:
    try:
        return ET.fromstring(xml_text)
    except ET.ParseError:
        return None


# ── EcoCyc API calls ───────────────────────────────────────────────────────────
def _ecocyc_get(frameid: str) -> ET.Element | None:
    """Fetch a BioCyc frame by ID → ptools-XML root. Only frame IDs work (not gene names)."""
    url = f"{_ECOCYC_BASE}/getxml?{ECOCYC_ORG}:{frameid}"
    try:
        r = _SESSION.get(url, timeout=30)
        time.sleep(_DELAY["ecocyc"])
    except Exception as exc:
        print(f"\n  [EcoCyc] network error for {frameid!r}: {exc}", file=sys.stderr)
        return None
    if r.status_code != 200:
        print(f"\n  [EcoCyc] HTTP {r.status_code} for {frameid!r}  (head: {r.text[:120]!r})",
              file=sys.stderr)
        return None
    root = _parse(r.text)
    if root is None:
        print(f"\n  [EcoCyc] non-XML for {frameid!r}  (head: {r.text[:120]!r})", file=sys.stderr)
    return root


def _ecocyc_resolve_uniprot(accession: str) -> str:
    """UniProt accession → BioCyc protein frame ID via the foreignid endpoint.

    Response format: tab-separated  "{DB}:{ID}\\t{count}\\t{frameid}"
    e.g. "UNIPROT:P0A9A6\\t1\\tEG10347-MONOMER"
    """
    try:
        r = _SESSION.get(
            f"{_ECOCYC_BASE}/foreignid",
            params={"ids": f"UNIPROT:{accession}", "orgid": ECOCYC_ORG},
            timeout=30,
        )
        time.sleep(_DELAY["ecocyc"])
    except Exception as exc:
        print(f"\n  [EcoCyc] foreignid error for {accession!r}: {exc}", file=sys.stderr)
        return ""
    if r.status_code != 200:
        print(f"\n  [EcoCyc] foreignid HTTP {r.status_code} for {accession!r}", file=sys.stderr)
        return ""
    for line in r.text.strip().splitlines():
        parts = line.split("\t")
        if len(parts) >= 3:
            return parts[2].strip()
    return ""


# ── EcoCyc pathway class resolution ───────────────────────────────────────────
_PW_CLASS_CACHE: dict[str, str] = {}


def _pathway_class(
    pathway_frameid: str,
    _root: ET.Element | None = None,
    _visited: set[str] | None = None,
) -> str:
    """Walk the super-pathway chain; return first non-generic intermediate name.

    Accepts an already-fetched root to avoid re-fetching the same frame.
    Results are cached globally since pathways are shared across genes.
    """
    if pathway_frameid in _PW_CLASS_CACHE:
        return _PW_CLASS_CACHE[pathway_frameid]
    if _visited is None:
        _visited = set()
    if pathway_frameid in _visited or len(_visited) >= 3:
        return ""
    _visited.add(pathway_frameid)

    root = _root if _root is not None else _ecocyc_get(pathway_frameid)
    if root is None:
        return ""
    pw_el = _find(root, "Pathway", "pathway")
    if pw_el is None:
        return ""

    own_name = _text(pw_el, "common-name")
    super_frameids = [
        child.get("frameid", "")
        for tag in ("super-pathways", "in-pathway")
        for container in [_find(pw_el, tag)] if container is not None
        for child in container
        if child.get("frameid", "") and child.get("frameid", "") not in _visited
    ]

    if not super_frameids or own_name.lower() not in _GENERIC_PATHWAY_ROOTS:
        result = own_name
    else:
        up = _pathway_class(super_frameids[0], _visited=_visited)
        result = up if up else own_name

    _PW_CLASS_CACHE[pathway_frameid] = result
    return result


# ── EcoCyc fetch stage ────────────────────────────────────────────────────────
def _go_classify(gene: Gene, go_id: str, go_name: str, aspect: str) -> None:
    """Route a GO annotation into the correct aspect dict and vote counter."""
    a = aspect.lower()
    if "biological" in a or a == "p":
        gene.go_bp.setdefault(go_id, go_name)
        gene.go_bp_votes[go_id] += 1
    elif "molecular" in a or a == "f":
        gene.go_mf.setdefault(go_id, go_name)
    elif "cellular" in a or "component" in a or a == "c":
        gene.go_cc.setdefault(go_id, go_name)


def fetch_ecocyc(gene: Gene) -> None:
    if not gene.accession:
        return
    protein_frameid = _ecocyc_resolve_uniprot(gene.accession)
    if not protein_frameid:
        return
    gene.ecocyc_id = protein_frameid

    root = _ecocyc_get(protein_frameid)
    if root is None:
        return

    # Navigate from protein frame → linked gene frame for full_name
    gene_frameid = ""
    for el in root.iter():
        if _local(el.tag) == "gene":
            for child in el:
                fid = child.get("frameid", "")
                if fid:
                    gene_frameid = fid
                    break
        if gene_frameid:
            break
    if gene_frameid and not gene.full_name:
        gene_root = _ecocyc_get(gene_frameid)
        if gene_root is not None:
            gene.full_name = _text(gene_root, "common-name")

    # GO annotations (two XML representations used by EcoCyc)
    for ann in _findall(root, "go-annotation"):
        go_id = _text(ann, "go-id")
        if go_id:
            _go_classify(gene, go_id, _text(ann, "go-term", "common-name"),
                         _text(ann, "aspect", "go-aspect"))
    for gt in _findall(root, "go-term"):
        go_id = gt.get("frameid", "")
        if go_id.startswith("GO:"):
            _go_classify(gene, go_id, _text(gt, "common-name"),
                         _text(gt, "aspect", "go-aspect"))

    # Pathway membership → ecocyc_class via super-pathway hierarchy
    pw_frameids = [
        child.get("frameid", "")
        for tag in ("catalyzes", "in-pathway", "pathways")
        for container in [_find(root, tag)] if container is not None
        for child in container
        if child.get("frameid", "")
    ]
    classes: list[str] = []
    for fid in pw_frameids[:6]:
        pw_root = _ecocyc_get(fid)
        if pw_root is None:
            continue
        pw_el = _find(pw_root, "Pathway")
        if pw_el is None:
            continue
        pw_name = _text(pw_el, "common-name")
        if pw_name:
            gene.ecocyc_pathways[fid] = pw_name
        cls = _pathway_class(fid, _root=pw_root)
        if cls:
            classes.append(cls)
    if classes:
        gene.ecocyc_class = Counter(classes).most_common(1)[0][0]

    # Molecular weight
    mw_text = _text(root, "molecular-weight-kd")
    if mw_text:
        try:
            gene.molecular_weight_kd = float(mw_text)
        except ValueError:
            pass


# ── UniProt fetch stage (secondary) ───────────────────────────────────────────
def fetch_uniprot(gene: Gene) -> None:
    try:
        results = _http_get(
            f"{_UNIPROT_BASE}/search",
            params={
                "query":  f"gene_exact:{gene.name} AND organism_id:{ORGANISM_TAXID} AND reviewed:true",
                "format": "json",
                "fields": "accession,protein_name,go,xref_kegg",
                "size":   "1",
            },
            source="uniprot",
        ).json().get("results", [])
    except Exception as exc:
        print(f"\n  [UniProt] error for {gene.name!r}: {exc}", file=sys.stderr)
        return
    if not results:
        return
    data = results[0]

    if not gene.accession:
        gene.accession = data.get("primaryAccession", "")
    if not gene.full_name:
        pd = data.get("proteinDescription", {})
        rec = pd.get("recommendedName") or next(iter(pd.get("submissionNames", [{}])), {})
        gene.full_name = rec.get("fullName", {}).get("value", "") if rec else ""

    for xref in data.get("uniProtKBCrossReferences", []):
        db = xref.get("database", "")
        if db == "KEGG" and not gene.kegg_id:
            gene.kegg_id = xref.get("id", "")
        elif db == "GO":
            props = {p["key"]: p["value"] for p in xref.get("properties", [])}
            term  = props.get("GoTerm", "")
            if len(term) >= 3:
                _go_classify(gene, xref["id"], term[2:].strip(), term[0])


# ── KEGG fetch stage (tertiary) ───────────────────────────────────────────────
def fetch_kegg(gene: Gene) -> None:
    kid = gene.kegg_id
    if not kid:
        try:
            lines = _http_get(f"{_KEGG_BASE}/find/{ORGANISM_KEGG}/{gene.name}",
                               source="kegg").text.strip().splitlines()
            kid = lines[0].split("\t")[0] if lines else ""
        except Exception:
            return
    if not kid:
        return
    gene.kegg_id = kid

    try:
        pids = [
            line.split("\t")[1].replace("path:", "")
            for line in _http_get(f"{_KEGG_BASE}/link/pathway/{kid}", source="kegg")
            .text.strip().splitlines()
            if "\t" in line
            and line.split("\t")[1].startswith("path:")
            and line.split("\t")[1].replace("path:", "") not in _KEGG_GENERIC
        ]
    except Exception:
        return

    for pid in pids:
        try:
            raw = _http_get(f"{_KEGG_BASE}/list/{pid}", source="kegg").text.strip()
            gene.kegg_pathways[pid] = re.sub(
                rf"\s*-\s*{re.escape(ORGANISM_FULL.split()[0])}.*$",
                "", raw.split("\t")[-1],
            ).strip()
        except Exception:
            gene.kegg_pathways[pid] = pid


# ── Annotation pipeline ────────────────────────────────────────────────────────
def annotate_all() -> list[Gene]:
    genes  = [Gene(name=n) for n in GENE_NAMES]
    failed: list[str] = []

    for i, g in enumerate(genes, 1):
        print(f"[{i:2d}/{len(genes)}]  {g.name:<8}", end=" ", flush=True)
        fetch_uniprot(g)   # accession required before EcoCyc lookup
        fetch_ecocyc(g)
        fetch_kegg(g)

        # Grouping: EcoCyc pathway class → GO-BP majority vote → Unclassified
        go_bp_best = ""
        if g.go_bp_votes:
            best_goid  = g.go_bp_votes.most_common(1)[0][0]
            go_bp_best = g.go_bp.get(best_goid, "")
        g.group = g.ecocyc_class or go_bp_best or "Unclassified"

        ok = bool(g.ecocyc_id or g.accession)
        if not ok:
            failed.append(g.name)
        print(
            f"[{'OK  ' if ok else 'WARN'}]  ecocyc={g.ecocyc_id or '?':<14}"
            f"  group={g.group[:32]:<32}  GO-BP={len(g.go_bp)}"
        )

    if failed:
        print(f"\n  Warning: could not resolve: {', '.join(failed)}")
    return genes


# ── Build hierarchy ────────────────────────────────────────────────────────────
def build_hierarchy(genes: list[Gene]) -> dict[str, list[str]]:
    raw: dict[str, list[str]] = {}
    for g in genes:
        raw.setdefault(g.group, []).append(g.name)
    unclassified = raw.pop("Unclassified", [])
    ordered = dict(sorted(raw.items(), key=lambda kv: -len(kv[1])))
    if unclassified:
        ordered["Unclassified"] = unclassified
    return ordered


# ── ASCII tree ─────────────────────────────────────────────────────────────────
def print_tree(hier: dict[str, list[str]], genes: list[Gene], path: str) -> None:
    gmap = {g.name: g for g in genes}
    lines: list[str] = []
    n_genes, n_groups = sum(len(v) for v in hier.values()), len(hier)
    lines.append(f"{ORGANISM_SHORT}   ({n_genes} genes, {n_groups} groups)")
    lines.append("")

    groups = list(hier.items())
    for gi, (grp, gene_names) in enumerate(groups):
        last_grp = gi == len(groups) - 1
        lines.append(("└── " if last_grp else "├── ") + grp)
        vert = "    " if last_grp else "│   "
        for ki, name in enumerate(gene_names):
            g = gmap.get(name)
            label = f"{name:<6}  {g.full_name}" if g else name
            lines.append(vert + ("└── " if ki == len(gene_names) - 1 else "├── ") + label)
        if not last_grp:
            lines.append(vert)

    text = "\n".join(lines)
    print("\n" + text + "\n")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text + "\n")
    print(f"Tree saved  -> {path}")


# ── Excel export ───────────────────────────────────────────────────────────────
def export_excel(genes: list[Gene], hier: dict[str, list[str]], path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        sys.exit("pip install openpyxl")

    colors = {grp: _PALETTE[i % len(_PALETTE)] for i, grp in enumerate(hier)}

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Gene Classification"

    headers = [
        "Gene", "Accession", "Protein Name", "Group",
        "EcoCyc Pathways", "KEGG Pathways",
        "GO Biological Process", "GO Molecular Function", "GO Cellular Component",
        "Mol. Weight (kDa)",
    ]
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for row, g in enumerate(genes, 2):
        fill = PatternFill("solid", fgColor=colors.get(g.group, "FFFFFF"))
        mw   = f"{g.molecular_weight_kd:.1f}" if g.molecular_weight_kd is not None else ""
        for col, val in enumerate([
            g.name, g.accession, g.full_name, g.group,
            "; ".join(g.ecocyc_pathways.values()),
            "; ".join(g.kegg_pathways.values()),
            "; ".join(g.go_bp.values()),
            "; ".join(g.go_mf.values()),
            "; ".join(g.go_cc.values()),
            mw,
        ], 1):
            ws.cell(row, col, val).fill = fill

    col_widths = [8, 12, 42, 38, 65, 65, 65, 55, 55, 18]
    for col_letter, width in zip("ABCDEFGHIJ", col_widths):
        ws.column_dimensions[col_letter].width = width

    wb.save(path)
    print(f"Excel saved -> {path}")


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    here = os.path.dirname(os.path.abspath(__file__))
    print(f"\nFetching {len(GENE_NAMES)} genes from EcoCyc + UniProt + KEGG ...\n")
    login_ecocyc()
    print()
    genes = annotate_all()
    hier  = build_hierarchy(genes)
    print_tree(hier, genes, os.path.join(here, "gene_hierarchy.txt"))
    export_excel(genes, hier, os.path.join(here, "gene_classification.xlsx"))
    print("\nDone.")


if __name__ == "__main__":
    main()

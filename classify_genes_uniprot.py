#!/usr/bin/env python3
"""
classify_genes_uniprot.py — E. coli Gene Classifier (UniProt only)
Organism  : Escherichia coli K-12 MG1655
API       : UniProt REST  https://rest.uniprot.org  (reviewed entries only)

No authentication required.  Grouping is derived from the GO Biological
Process term with the highest annotation count across UniProt evidence codes.
Only gene names are hardcoded; everything else is resolved at runtime.
"""
from __future__ import annotations
import os, re, sys, time
from collections import Counter
from dataclasses import dataclass, field

sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]

try:
    import requests
except ImportError:
    sys.exit("pip install requests")

# ── Organism constants ─────────────────────────────────────────────────────────
ORGANISM_SHORT = "E. coli K-12"
ORGANISM_TAXID = "83333"

# ── Gene list — only names, nothing else ──────────────────────────────────────
GENE_NAMES: list[str] = [
    "ftsZ", "ftsI", "mrdA", "mrcB", "mrcA", "murA", "murC",
    "lpxA", "lpxC", "lptA", "lptC", "msbA",
    "folA", "folP", "dnaE", "dnaB", "gyrA", "gyrB", "parC", "parE",
    "rpoA", "rpoB", "rpsA", "rpsL", "rplA", "rplC", "secA", "secY",
]

# ── Display palette — no biological meaning ────────────────────────────────────
_PALETTE = [
    "AED6F1", "A9DFBF", "A3E4D7", "F9E79F", "F5CBA7",
    "D2B4DE", "FAD7A0", "FADBD8", "D5DBDB", "ABEBC6",
    "F1948A", "85C1E9", "82E0AA", "F8C471", "C39BD3",
    "76D7C4", "F0B27A", "EC7063", "5DADE2", "45B39D",
]

# ── Data model ─────────────────────────────────────────────────────────────────
@dataclass
class Gene:
    name:        str
    accession:   str = ""
    full_name:   str = ""
    go_bp:       dict[str, str] = field(default_factory=dict)  # GO-ID → term
    go_bp_votes: Counter = field(default_factory=Counter)       # GO-ID → count
    go_mf:       dict[str, str] = field(default_factory=dict)
    go_cc:       dict[str, str] = field(default_factory=dict)
    group:       str = "Unclassified"

# ── HTTP helper ────────────────────────────────────────────────────────────────
_UNIPROT_BASE = "https://rest.uniprot.org/uniprotkb"
_DELAY        = 0.15   # seconds between requests

_SESSION = requests.Session()
_SESSION.headers["User-Agent"] = "classify_genes_uniprot.py/1.0 (research script)"


def _go_classify(gene: Gene, go_id: str, go_name: str, aspect: str) -> None:
    """Route a GO cross-reference into the correct aspect dict and vote counter."""
    a = aspect.lower()
    if "biological" in a or a == "p":
        gene.go_bp.setdefault(go_id, go_name)
        gene.go_bp_votes[go_id] += 1
    elif "molecular" in a or a == "f":
        gene.go_mf.setdefault(go_id, go_name)
    elif "cellular" in a or "component" in a or a == "c":
        gene.go_cc.setdefault(go_id, go_name)


def fetch_uniprot(gene: Gene) -> None:
    try:
        results = _SESSION.get(
            f"{_UNIPROT_BASE}/search",
            params={
                "query":  f"gene_exact:{gene.name} AND organism_id:{ORGANISM_TAXID} AND reviewed:true",
                "format": "json",
                "fields": "accession,protein_name,go",
                "size":   "1",
            },
            timeout=30,
        ).json().get("results", [])
        time.sleep(_DELAY)
    except Exception as exc:
        print(f"\n  [UniProt] error for {gene.name!r}: {exc}", file=sys.stderr)
        return
    if not results:
        return
    data = results[0]

    gene.accession = data.get("primaryAccession", "")

    pd  = data.get("proteinDescription", {})
    rec = pd.get("recommendedName") or next(iter(pd.get("submissionNames", [{}])), {})
    gene.full_name = rec.get("fullName", {}).get("value", "") if rec else ""

    # Each GO ID may appear multiple times (one per evidence code).
    # Deduplicate before voting so each evidence code counts at most once.
    go_terms: dict[str, tuple[str, str]] = {}  # go_id → (name, aspect)
    for xref in data.get("uniProtKBCrossReferences", []):
        if xref.get("database") != "GO":
            continue
        props = {p["key"]: p["value"] for p in xref.get("properties", [])}
        term  = props.get("GoTerm", "")
        goid  = xref["id"]
        if len(term) >= 3 and (goid not in go_terms or not go_terms[goid][0]):
            go_terms[goid] = (term[2:].strip(), term[0])

    for go_id, (go_name, aspect) in go_terms.items():
        _go_classify(gene, go_id, go_name, aspect)


# ── Annotation pipeline ────────────────────────────────────────────────────────
def annotate_all() -> list[Gene]:
    genes  = [Gene(name=n) for n in GENE_NAMES]
    failed: list[str] = []

    for i, g in enumerate(genes, 1):
        print(f"[{i:2d}/{len(genes)}]  {g.name:<8}", end=" ", flush=True)
        fetch_uniprot(g)

        if g.go_bp_votes:
            best = g.go_bp_votes.most_common(1)[0][0]
            g.group = g.go_bp.get(best, "Unclassified")
        else:
            g.group = "Unclassified"

        ok = bool(g.accession)
        if not ok:
            failed.append(g.name)
        print(
            f"[{'OK  ' if ok else 'WARN'}]  acc={g.accession:<12}"
            f"  group={g.group[:36]:<36}  GO-BP={len(g.go_bp)}"
        )

    if failed:
        print(f"\n  Warning: could not resolve: {', '.join(failed)}")
    return genes


# ── Build hierarchy ────────────────────────────────────────────────────────────
def build_hierarchy(genes: list[Gene]) -> dict[str, list[str]]:
    raw: dict[str, list[str]] = {}
    for g in genes:
        raw.setdefault(g.group, []).append(g.name)
    unclassified = raw.pop("Unclassified", [])
    ordered = dict(sorted(raw.items(), key=lambda kv: -len(kv[1])))
    if unclassified:
        ordered["Unclassified"] = unclassified
    return ordered


# ── ASCII tree ─────────────────────────────────────────────────────────────────
def print_tree(hier: dict[str, list[str]], genes: list[Gene], path: str) -> None:
    gmap = {g.name: g for g in genes}
    lines: list[str] = []
    n_genes, n_groups = sum(len(v) for v in hier.values()), len(hier)
    lines.append(f"{ORGANISM_SHORT}   ({n_genes} genes, {n_groups} groups)")
    lines.append("")

    groups = list(hier.items())
    for gi, (grp, gene_names) in enumerate(groups):
        last_grp = gi == len(groups) - 1
        lines.append(("└── " if last_grp else "├── ") + grp)
        vert = "    " if last_grp else "│   "
        for ki, name in enumerate(gene_names):
            g = gmap.get(name)
            label = f"{name:<6}  {g.full_name}" if g else name
            lines.append(vert + ("└── " if ki == len(gene_names) - 1 else "├── ") + label)
        if not last_grp:
            lines.append(vert)

    text = "\n".join(lines)
    print("\n" + text + "\n")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(text + "\n")
    print(f"Tree saved  -> {path}")


# ── Excel export ───────────────────────────────────────────────────────────────
def export_excel(genes: list[Gene], hier: dict[str, list[str]], path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        sys.exit("pip install openpyxl")

    colors = {grp: _PALETTE[i % len(_PALETTE)] for i, grp in enumerate(hier)}

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Gene Classification"

    headers = [
        "Gene", "Accession", "Protein Name", "Group",
        "GO Biological Process", "GO Molecular Function", "GO Cellular Component",
    ]
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for row, g in enumerate(genes, 2):
        fill = PatternFill("solid", fgColor=colors.get(g.group, "FFFFFF"))
        for col, val in enumerate([
            g.name, g.accession, g.full_name, g.group,
            "; ".join(g.go_bp.values()),
            "; ".join(g.go_mf.values()),
            "; ".join(g.go_cc.values()),
        ], 1):
            ws.cell(row, col, val).fill = fill

    col_widths = [8, 12, 42, 38, 65, 55, 55]
    for col_letter, width in zip("ABCDEFG", col_widths):
        ws.column_dimensions[col_letter].width = width

    wb.save(path)
    print(f"Excel saved -> {path}")


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    here = os.path.dirname(os.path.abspath(__file__))
    print(f"\nFetching {len(GENE_NAMES)} genes from UniProt ...\n")
    genes = annotate_all()
    hier  = build_hierarchy(genes)
    print_tree(hier, genes, os.path.join(here, "gene_hierarchy_uniprot.txt"))
    export_excel(genes, hier, os.path.join(here, "gene_classification_uniprot.xlsx"))
    print("\nDone.")


if __name__ == "__main__":
    main()
